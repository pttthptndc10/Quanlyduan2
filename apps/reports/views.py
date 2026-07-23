"""
Reports Views — Thống kê, Quản lý Gom hàng linh kiện & Xuất file Excel (.xlsx).
"""

import json
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from apps.authentication.services.supabase_client import get_admin_supabase_client
from core.utils import get_user_from_session, TASK_STATUS_LABELS, PRIORITY_LABELS


def reports_view(request):
    user = get_user_from_session(request)
    db = get_admin_supabase_client(request)
    try:
        tasks    = db.table('tasks').select('*').execute().data or []
        projects = db.table('projects').select('*').execute().data or []
        members  = db.table('profiles').select('id, full_name, role').eq('status', 'active').execute().data or []
        today    = date.today().isoformat()

        # Task by status
        by_status = {s: 0 for s in TASK_STATUS_LABELS}
        for t in tasks:
            s = t.get('status', 'todo')
            by_status[s] = by_status.get(s, 0) + 1

        # Tasks per member
        member_stats = []
        for m in members:
            mt = [t for t in tasks if t.get('assignee_id') == m['id']]
            member_stats.append({
                'name': m['full_name'],
                'total': len(mt),
                'done': sum(1 for t in mt if t['status'] == 'done'),
                'overdue': sum(1 for t in mt if t.get('deadline') and t['deadline'] < today and t['status'] not in ('done', 'cancelled')),
            })

        # Component Batches (Phiên gom hàng linh kiện)
        batches = db.table('component_batches').select('*, profiles!component_batches_created_by_fkey(full_name)') \
            .order('created_at', desc=True).execute().data or []

    except Exception:
        by_status = {}
        projects = []
        member_stats = []
        batches = []

    return render(request, 'reports/reports.html', {
        'by_status': by_status,
        'status_labels': TASK_STATUS_LABELS,
        'member_stats': member_stats,
        'total_projects': len(projects),
        'batches': batches,
        'user': user,
    })


def create_batch(request):
    """Tạo phiên gom hàng linh kiện mới."""
    if request.method != 'POST':
        return redirect('/reports/')
    user = get_user_from_session(request)
    name = request.POST.get('name', '').strip()
    if name:
        db = get_admin_supabase_client(request)
        db.table('component_batches').insert({
            'name': name,
            'status': 'active',
            'created_by': user['id'],
        }).execute()
    return redirect('/reports/')


def export_batch_excel(request, batch_id):
    """Gộp & Xuất file Excel (.xlsx) tổng hợp linh kiện từ phiên gom hàng."""
    db = get_admin_supabase_client(request)
    try:
        batch = db.table('component_batches').select('*').eq('id', batch_id).single().execute().data
        files = db.table('component_files').select('*, projects(name), profiles!component_files_created_by_fkey(full_name)') \
            .eq('batch_id', batch_id).execute().data or []
    except Exception:
        return redirect('/reports/')

    # Tạo Workbook Excel bằng openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gom Linh Kien"

    # Style header
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    headers = ['STT', 'Dự án', 'Người nhập', 'Tên linh kiện', 'Số lượng', 'Đơn giá (VND)', 'Thành tiền (VND)', 'Link Shop', 'Ghi chú']
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    row_idx = 1
    total_amount = 0

    for file_obj in files:
        project_name = (file_obj.get('projects') or {}).get('name', 'Không rõ dự án')
        creator_name = (file_obj.get('profiles') or {}).get('full_name', 'Không rõ')
        items = file_obj.get('content') or []
        if isinstance(items, str):
            try:
                items = json.loads(items)
            except Exception:
                items = []

        for item in items:
            qty = int(item.get('quantity', 1))
            price = float(item.get('price', 0))
            line_total = qty * price
            total_amount += line_total

            row_data = [
                row_idx,
                project_name,
                creator_name,
                item.get('name', ''),
                qty,
                price,
                line_total,
                item.get('shop', ''),
                item.get('notes', ''),
            ]
            ws.append(row_data)

            current_row = ws.max_row
            ws.cell(row=current_row, column=1).alignment = align_center
            ws.cell(row=current_row, column=2).alignment = align_left
            ws.cell(row=current_row, column=3).alignment = align_left
            ws.cell(row=current_row, column=4).alignment = align_left
            ws.cell(row=current_row, column=5).alignment = align_center
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7).number_format = '#,##0'

            for c in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=c).border = thin_border

            row_idx += 1

    # Dòng tổng cộng
    total_row = ['TỔNG CỘNG', '', '', '', '', '', total_amount, '', '']
    ws.append(total_row)
    last_row = ws.max_row
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
    total_cell = ws.cell(row=last_row, column=1)
    total_cell.font = Font(name="Arial", size=11, bold=True, color="0891B2")
    total_cell.alignment = Alignment(horizontal="right", vertical="center")

    total_val_cell = ws.cell(row=last_row, column=7)
    total_val_cell.font = Font(name="Arial", size=11, bold=True, color="0891B2")
    total_val_cell.number_format = '#,##0'

    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Xuất response file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"GomLinhKien_{(batch.get('name') or 'Batch').replace(' ', '_')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_tasks_csv(request):
    """Export toàn bộ tasks ra CSV."""
    db = get_admin_supabase_client(request)
    tasks = db.table('tasks').select('*, projects(name)').execute().data or []
    profiles = db.table('profiles').select('id, full_name').execute().data or []
    prof_map = {p['id']: p['full_name'] for p in profiles}

    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="tasks_export.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['Tiêu đề', 'Dự án', 'Người thực hiện', 'Trạng thái', 'Ưu tiên', 'Tiến độ %', 'Deadline'])
    for t in tasks:
        writer.writerow([
            t.get('title', ''),
            (t.get('projects') or {}).get('name', ''),
            prof_map.get(t.get('assignee_id'), '—'),
            TASK_STATUS_LABELS.get(t.get('status', ''), t.get('status', '')),
            PRIORITY_LABELS.get(t.get('priority', ''), t.get('priority', '')),
            t.get('progress', 0),
            t.get('deadline', ''),
        ])
    return response
